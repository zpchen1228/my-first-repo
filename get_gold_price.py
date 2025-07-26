import os
import requests
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
import schedule


def download_gold_price():
    """从上海黄金交易所下载每日黄金价格并保存到Excel"""
    # 文件路径设置
    save_dir = r"D:\czp\data\auto_download\gold_price"
    file_path = os.path.join(save_dir, "gold_price.xlsx")

    # 确保目录存在
    os.makedirs(save_dir, exist_ok=True)

    # 1. 检查文件是否存在并读取现有数据的最新日期
    existing_date = None
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        last_row = ws.max_row
        if last_row > 1:  # 跳过标题行
            existing_date = ws.cell(last_row, 3).value  # 日期在第三列

    # 2. 获取当前日期并检查是否最新
    today = datetime.now().strftime("%Y-%m-%d")
    if existing_date == today:
        print(f"{today} 数据已是最新，无需更新")
        return

    # 3. 构建目标URL（使用当天日期）
    base_url = "https://www.sge.com.cn/sjzx/quotation_daily_new"
    params = {
        "start_date": today,
        "end_date": today
    }

    try:
        # 4. 发送请求获取数据
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Referer": "https://www.sge.com.cn/sjzx/quotation_daily_new"
        }
        response = requests.get(base_url, params=params, headers=headers, timeout=10)
        response.raise_for_status()

        # 5. 解析HTML获取黄金价格
        soup = BeautifulSoup(response.text, 'html.parser')

        # 修复：使用更通用的表格定位方式
        table = None
        # 尝试多种可能的表格定位方式
        for class_name in ['table', 'data-list', 'data-table', 'list']:
            table = soup.find('table', class_=class_name)
            if table:
                break

        # 如果找不到表格，抛出有详细信息的异常
        if not table:
            print("页面HTML结构：", response.text[:1000])  # 输出前1000字符调试
            raise ValueError(f"找不到表格元素，请检查页面结构 - {today}")

        # 查找AU99.99的收盘价
        closing_price = None
        for row in table.find_all('tr')[1:]:  # 跳过表头
            columns = [col.text.strip() for col in row.find_all('td')]
            if len(columns) >= 3 and 'au99.99' in columns[0].lower():
                # 尝试不同列位置获取收盘价
                closing_price = columns[1] if len(columns) >= 2 else None
                if not closing_price and len(columns) >= 4:  # 如果第二列空，尝试第四列
                    closing_price = columns[3]
                break

        if not closing_price:
            # 尝试备用查找方法
            gold_row = soup.find(string=lambda text: text and 'au99.99' in text.lower())
            if gold_row and gold_row.parent and gold_row.parent.parent:
                columns = gold_row.parent.parent.find_all('td')
                if len(columns) >= 2:
                    closing_price = columns[1].text.strip()

            if not closing_price:
                raise ValueError("未找到AU99.99的收盘价数据")

        # 6. 准备Excel文件
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
            start_row = ws.max_row + 1  # 在现有数据后追加
        else:
            wb = Workbook()
            ws = wb.active
            start_row = 2
            # 添加标题行
            ws.cell(1, 1).value = "ID"
            ws.cell(1, 2).value = "Product"
            ws.cell(1, 3).value = "Date"
            ws.cell(1, 4).value = "Closing_Price"

        # 7. 写入新数据
        # 清理价格中的特殊字符
        clean_price = ''.join(filter(lambda x: x in '0123456789.', closing_price))
        unique_id = f"{today}-AU99.99-{int(time.time())}"
        ws.cell(start_row, 1).value = unique_id
        ws.cell(start_row, 2).value = "AU99.99"
        ws.cell(start_row, 3).value = today
        ws.cell(start_row, 4).value = float(clean_price)

        # 8. 保存文件
        wb.save(file_path)
        print(f"成功保存黄金价格 {clean_price} 到: {file_path}")

    except Exception as e:
        print(f"获取数据时出错: {e}")


def schedule_daily_download():
    """设置定时任务每天执行"""
    # 每天下午4点执行（交易所数据更新时间后）
    schedule.every().day.at("16:00").do(download_gold_price)

    print("定时任务已启动，每天下午4点自动下载数据...")
    while True:
        schedule.run_pending()
        time.sleep(60)  # 每分钟检查一次


if __name__ == "__main__":
    # 直接执行一次
    download_gold_price()

    # 如果要设置定时任务，取消下面行的注释
    # schedule_daily_download()
