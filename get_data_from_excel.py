from openpyxl import load_workbook


def get_latest_rates(file_path, currencies):
    """
    从Excel文件底部向上获取指定货币的最新汇率
    :param file_path: Excel文件路径
    :param currencies: 货币代号列表 ["USD", "EUR"]
    :return: {货币代号: 汇率值} 的字典
    """
    rates = {}
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        last_row = ws.max_row

        # 从最后一行向上遍历（跳过标题行）
        for row in range(last_row, 1, -1):
            currency_name = ws.cell(row, 2).value.strip()

            # 检查是否匹配目标货币
            for currency_code in currencies:
                if currency_code in currency_name:
                    rates[currency_code] = ws.cell(row, 3).value
                    # 如果已经找到所有货币，提前结束遍历
                    if len(rates) == len(currencies):
                        return rates
    except Exception as e:
        print(f"Error reading rates: {e}")

    return rates

