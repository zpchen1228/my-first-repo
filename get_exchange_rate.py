import os
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime


def download_exchange_rate_data():
    """从中国人民银行下载汇率数据并保存到Excel"""
    # 文件路径设置
    save_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(save_dir, "exchange_rate.xlsx")

    # 确保目录存在
    os.makedirs(save_dir, exist_ok=True)

    # 1. 检查文件是否存在并读取现有数据的最新日期
    existing_date = None
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        last_row = ws.max_row
        if last_row > 1:
            existing_date = ws.cell(last_row, 4).value

    # 2. 从API获取最新数据日期
    try:
        url = "https://www.chinamoney.com.cn/r/cms/www/chinamoney/data/fx/ccpr.json"
        res = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        res.raise_for_status()
        api_date = res.json()["data"]["lastDate"]

        # 3. 检查数据是否最新
        if existing_date == api_date:
            print("This data is the latest.")
            return
    except Exception as e:
        print(f"获取数据时出错: {e}")
        return

    # 4. 获取完整数据
    data = res.json()
    data_list = data["records"]

    # 5. 准备Excel文件
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        start_row = ws.max_row + 1  # 在现有数据后追加
    else:
        wb = Workbook()
        ws = wb.active
        start_row = 2
        # 首字母大写标题
        ws.cell(1, 1).value = "Id"
        ws.cell(1, 2).value = "Currency"
        ws.cell(1, 3).value = "Exchange Rate"
        ws.cell(1, 4).value = "Date"

    # 6. 追加新数据
    for idx, item in enumerate(data_list, start=start_row):
        # 生成唯一ID (日期+货币代码+时间戳)
        unique_id = f"{api_date}-{item['vrtCode']}-{datetime.now().strftime('%H%M%S%f')}"
        ws.cell(idx, 1).value = unique_id
        ws.cell(idx, 2).value = item["vrtEName"]
        ws.cell(idx, 3).value = item["price"]
        ws.cell(idx, 4).value = api_date

    # 7. 保存文件
    wb.save(file_path)
    print(f"成功保存{len(data_list)}条记录到: {file_path}")


if __name__ == "__main__":
    download_exchange_rate_data()